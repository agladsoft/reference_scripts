import os
import sys
import json
import warnings
import requests
import app_logger
import contextlib
import pandas as pd
from typing import Optional
from datetime import datetime
from dotenv import load_dotenv
from clickhouse_connect import get_client
from clickhouse_connect.driver import Client
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv()

DATE_FORMATS: list = [
    "%m/%d/%y",
    "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S",
    "%m/%d/%Y",
    "%d%b%Y"
]

list_join_columns: list = ["telephone_number", "email"]

logger: app_logger = app_logger.get_logger(os.path.basename(__file__).replace(".py", "_") + str(datetime.now().date()))

headers_eng: dict = {
    ("ИНН",): "inn",
    ("Наименование",): "company_name",
    ("КПП",): "kpp",
    ("ОГРН",): "ogrn",
    ("ФИО руководителя", "Ген.директор"): "director_full_name",
    ("Должность руководителя",): "position",
    ("Номер телефона", "Телефон"): list_join_columns[0],
    ("Дополнительный телефон 1",): list_join_columns[0],
    ("Дополнительный телефон 2",): list_join_columns[0],
    ("Дополнительный телефон 3",): list_join_columns[0],
    ("Дополнительный телефон 4",): list_join_columns[0],
    ("Дополнительный телефон 5",): list_join_columns[0],
    ("Дополнительный телефон 6",): list_join_columns[0],
    ("Дополнительный телефон 7",): list_join_columns[0],
    ("Дополнительный телефон 8",): list_join_columns[0],
    ("Дополнительный телефон 9",): list_join_columns[0],
    ("Электронная почта", "E-MAIL"): list_join_columns[1],
    ("Дополнительная электронная почта 1",): list_join_columns[1],
    ("Дополнительная электронная почта 2",): list_join_columns[1],
    ("Дополнительная электронная почта 3",): list_join_columns[1],
    ("Дополнительная электронная почта 4",): list_join_columns[1],
    ("Дополнительная электронная почта 5",): list_join_columns[1],
    ("Дополнительная электронная почта 6",): list_join_columns[1],
    ("Дополнительная электронная почта 7",): list_join_columns[1],
    ("Дополнительная электронная почта 8",): list_join_columns[1],
    ("Дополнительная электронная почта 9",): list_join_columns[1],
    ("Адрес",): "address",
    ("Регион по адресу",): "region",
    ("Ссылка на сайт",): "website_link",
    ("Карточка в Фокусе",): "link_to_card_in_focus",
    ("Статус",): "status_at_upload_date",
    ("Выручка, тыс. руб",): "revenue_at_upload_date_thousand_rubles",
    ("Чистая прибыль/ убыток, тыс. руб",): "net_profit_or_loss_at_upload_date_thousand_rubles",
    ("Количество сотрудников",): "employees_number_at_upload_date",
    ("Полученные лицензии",): "licenses",
    ("Дата регистрации",): "registration_date",
    ("Реестр МСП",): "register_msp",
    ("Основной вид деятельности",): "activity_main_type",
    ("Другие виды деятельности",): "activity_other_types",
    ("Регион регистрации",): "registration_region",
    ("Филиалы",): "branch_name"
}


def get_my_env_var(var_name: str) -> str:
    try:
        return os.environ[var_name]
    except KeyError as e:
        raise MissingEnvironmentVariable(f"{var_name} does not exist") from e


class MissingEnvironmentVariable(Exception):
    pass


class ReferenceCompass(object):
    def __init__(self, input_file_path: str, output_folder: str):
        self.table_name: str = "cache_dadata"
        self.input_file_path: str = input_file_path
        self.output_folder: str = output_folder
        self.original_columns: dict = {}

    @staticmethod
    def connect_to_db() -> Client:
        """
        Connecting to clickhouse.
        """
        try:
            client: Client = get_client(host=get_my_env_var('HOST'), database=get_my_env_var('DATABASE'),
                                        username=get_my_env_var('USERNAME_DB'), password=get_my_env_var('PASSWORD'))
            client.query("SET allow_experimental_lightweight_delete=1")
        except Exception as ex_connect:
            logger.error(f"Error connection to db {ex_connect}. Type error is {type(ex_connect)}.")
            print("error_connect_db", file=sys.stderr)
            sys.exit(1)
        return client

    def change_data_in_db(self, parsed_data: list) -> None:
        """
        Delete the data from the database if the row is loaded now.
        """
        client = self.connect_to_db()
        parsed_data_copy: list = parsed_data.copy()
        for dict_data in parsed_data_copy:
            for key, value in dict_data.items():
                if key in ["inn"]:
                    try:
                        for _ in client.query(f"SELECT * FROM reference_compass WHERE inn='{value}'").result_rows:
                            client.query(f"DELETE FROM reference_compass WHERE inn='{value}'")
                        break
                    except Exception as ex_db:
                        logger.error(f"Failed to execute action. Error is {ex_db}. Type error is {type(ex_db)}. "
                                     f"Data is {dict_data}")
                        self.save_to_csv(dict_data)

    @staticmethod
    def leave_largest_data_with_dupl_inn(parsed_data: list) -> list:
        """
        Leave the rows with the largest amount of data with repeated INN.
        """
        uniq_parsed_parsed_data: list = []
        for d in parsed_data:
            if [cache for cache in uniq_parsed_parsed_data if d["inn"] == cache["inn"]]:
                index_dupl: int = uniq_parsed_parsed_data.index(next(filter(lambda n: n.get('inn') == d["inn"],
                                                                            uniq_parsed_parsed_data)))
                if len([x for x in list(d.values()) if x is not None]) > \
                        len([x for x in uniq_parsed_parsed_data[index_dupl].values() if x is not None]):
                    uniq_parsed_parsed_data.pop(index_dupl)
                    uniq_parsed_parsed_data.append(d)
            else:
                uniq_parsed_parsed_data.append(d)
        return uniq_parsed_parsed_data

    def get_data_from_cache(self, dict_data: dict, index: int):
        """
        Get data from the cache in order not to go to dadata again, because the limit is 10000 requests per day.
        """
        self.get_data_from_service_inn(dict_data, index)
        if not dict_data["dadata_branch_name"] \
                and not dict_data["dadata_branch_address"] and not dict_data["dadata_branch_region"]:
            dict_data["dadata_branch_name"] = None
            dict_data["dadata_branch_address"] = None
            dict_data["dadata_branch_region"] = None

    @staticmethod
    def convert_format_date(date: str) -> Optional[str]:
        """
        Convert to a date type.
        """
        for date_format in DATE_FORMATS:
            with contextlib.suppress(ValueError):
                return str(datetime.strptime(date, date_format).date())
        return None

    def handle_raw_data(self, parsed_data: list) -> None:
        """
        Change data types or changing values.
        """
        for index, dict_data in enumerate(parsed_data, 2):
            for key, value in dict_data.items():
                with contextlib.suppress(Exception):
                    if key in ["registration_date"]:
                        dict_data[key] = self.convert_format_date(value) if value else None
                    elif key in ["revenue_at_upload_date_thousand_rubles", "employees_number_at_upload_date",
                                 "net_profit_or_loss_at_upload_date_thousand_rubles"]:
                        dict_data[key] = int(value) if value.isdigit() else None
                    elif value == 'None':
                        dict_data[key] = None
            self.add_new_columns(dict_data)
            self.get_data_from_cache(dict_data, index)

    def add_new_columns(self, dict_data: dict) -> None:
        """
        Add new columns.
        """
        dict_data['original_file_name'] = os.path.basename(self.input_file_path)
        dict_data['original_file_parsed_on'] = str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        dict_data["dadata_branch_name"] = ''
        dict_data["dadata_branch_address"] = ''
        dict_data["dadata_branch_region"] = ''

    @staticmethod
    def add_dadata_columns(company_data: dict, company_address: dict, company_address_data: dict,
                           company_data_branch: dict, company: dict, dict_data: dict,
                           is_company_name_from_cache: bool) -> None:
        """
        Add values from dadata to the dictionary.
        """
        dict_data["dadata_company_name"] = \
            f'{company_data.get("opf").get("short", "") if company_data.get("opf") else ""} ' \
            f'{company_data["name"]["full"]}'.strip() \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_company_name"]
        dict_data["dadata_address"] = company_address.get("unrestricted_value") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_address"]
        dict_data["dadata_region"] = company_address_data.get("region_with_type") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_region"]
        dict_data["dadata_federal_district"] = company_address_data.get("federal_district") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_federal_district"]
        dict_data["dadata_city"] = company_address_data.get("city") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_city"]
        dict_data["dadata_okved_activity_main_type"] = company_data.get("okved") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_okved_activity_main_type"]
        dict_data["dadata_branch_name"] += f'{company.get("value")}, КПП {company_data.get("kpp", "")}' + '\n' \
            if company_data_branch == "BRANCH" else ''
        dict_data["dadata_branch_address"] += company_address["unrestricted_value"] + '\n' \
            if company_data_branch == "BRANCH" else ''
        dict_data["dadata_branch_region"] += company_address_data["region_with_type"] + '\n' \
            if company_data_branch == "BRANCH" else ''
        dict_data["dadata_geo_lat"] = company_address_data.get("geo_lat") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_geo_lat"]
        dict_data["dadata_geo_lon"] = company_address_data.get("geo_lon") \
            if company_data_branch == "MAIN" or not company_data_branch else dict_data["dadata_geo_lat"]
        dict_data["is_company_name_from_cache"] = is_company_name_from_cache

    def get_data_from_dadata(self, dadata_request: list, dict_data: dict, index: int) -> None:
        """
        Get data from dadata.
        """
        for company in dadata_request[0]:
            try:
                company_data: dict = company.get("data")
                company_address: dict = company_data.get("address") or {}
                company_address_data: dict = company_address.get("data", {})
                company_data_branch: dict = company_data.get("branch_type")
                dict_data["dadata_status"] = company_data["state"]["status"]
                dict_data["dadata_registration_date"] = \
                    datetime.utcfromtimestamp(company_data["state"]["registration_date"] // 1000).strftime('%Y-%m-%d') \
                    if company_data["state"]["registration_date"] else None
                dict_data["dadata_liquidation_date"] = \
                    datetime.utcfromtimestamp(company_data["state"]["liquidation_date"] // 1000).strftime('%Y-%m-%d') \
                    if company_data["state"]["liquidation_date"] else None
                if company_data and company_data["state"]["status"] != "LIQUIDATED":
                    self.add_dadata_columns(company_data, company_address, company_address_data, company_data_branch,
                                            company, dict_data, dadata_request[1])
            except Exception as ex_parse:
                logger.error(f"Error code: error processing in row {index + 1}! "
                             f"Error is {ex_parse} Data is {dict_data}")
                self.save_to_csv(dict_data)

    def get_data_from_service_inn(self, dict_data: dict, index: int) -> None:
        """
        Connect to dadata.
        """
        data: dict = {
            "inn": dict_data["inn"]
        }
        response = requests.post("http://service_inn:8003", json=data)
        if response.status_code == 200:
            self.get_data_from_dadata(response.json(), dict_data, index)

    def save_to_csv(self, dict_data: dict) -> None:
        df: pd.DataFrame = pd.DataFrame([dict_data])
        index_of_column: int = df.columns.get_loc('original_file_name')
        columns_slice: pd.DataFrame = df.iloc[:, :index_of_column]
        columns_slice.rename(columns=self.original_columns, inplace=True)
        with open(f"{os.path.dirname(self.input_file_path)}/{os.path.basename(self.input_file_path)}_errors.csv", 'a') \
                as f:
            columns_slice.to_csv(f, header=f.tell() == 0, index=False)

    def write_to_json(self, parsed_data: list) -> None:
        """
        Write data to json.
        """
        basename: str = os.path.basename(self.input_file_path)
        output_file_path: str = os.path.join(self.output_folder, f'{basename}.json')
        with open(f"{output_file_path}", 'w', encoding='utf-8') as f:
            json.dump(parsed_data, f, ensure_ascii=False, indent=4)

    def get_column_eng(self, column: tuple, dict_header: dict) -> None:
        """
        Get the English column name.
        """
        for cell in column:
            for key, value in headers_eng.items():
                for column_rus in key:
                    if cell.internal_value == column_rus:
                        self.original_columns[value] = cell.internal_value
                        dict_header[cell.column_letter] = cell.internal_value, value

    @staticmethod
    def get_value_from_cell(index: int, column: tuple, dict_header: dict, dict_columns: dict) -> None:
        """
        Get a value from a cell, including url.
        """
        for cell in column:
            for key, value in dict_header.items():
                if cell.column_letter == key:
                    try:
                        if value[1] in list_join_columns and dict_columns.get(value[1]):
                            dict_columns[value[1]] = f"{dict_columns.get(value[1])}/{cell.value}"
                            continue
                        dict_columns[value[1]] = cell.hyperlink.target
                    except AttributeError:
                        if value[1] == 'inn' and (len(str(cell.value)) < 10 or len(str(cell.value)) == 11 or
                                                  len(str(cell.value)) > 12):
                            logger.error(f"Error code: error processing in row {index + 1}!")
                            cell.value = f"0{cell.value}"
                            # print(f"in_row_{index + 1}", file=sys.stderr)
                            # sys.exit(1)
                        dict_columns[value[1]] = str(cell.value)

    def parse_xlsx(self, ws: Worksheet, parsed_data: list) -> None:
        """
        Xlsx file parsing.
        """
        dict_header: dict = {}
        for i, column in enumerate(ws):
            dict_columns: dict = {}
            if i == 0:
                self.get_column_eng(column, dict_header)
                continue
            self.get_value_from_cell(i, column, dict_header, dict_columns)
            parsed_data.append(dict_columns)

    def main(self) -> None:
        """
        The main function where we read the Excel file and write the file to json.
        """
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            logger.info(f"Filename is {self.input_file_path}")
            wb: Workbook = load_workbook(self.input_file_path)
            ws: Worksheet = wb[wb.sheetnames[0]]
            parsed_data: list = []
            self.parse_xlsx(ws, parsed_data)
            self.handle_raw_data(parsed_data)
            parsed_data: list = self.leave_largest_data_with_dupl_inn(parsed_data)
            self.change_data_in_db(parsed_data)
            self.write_to_json(parsed_data)
            logger.info("The script has completed its work")


if __name__ == "__main__":
    reference_compass: ReferenceCompass = ReferenceCompass(sys.argv[1], sys.argv[2])
    try:
        reference_compass.main()
    except Exception as ex:
        logger.error(f"Error code: unknown error - {ex}!")
        print("unknown_error", file=sys.stderr)
        sys.exit(1)

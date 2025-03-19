import os
import sys
from unittest.mock import MagicMock
from unittest import mock
import pytest
from openpyxl import Workbook
from requests.exceptions import RequestException

# Импортируем тестируемый модуль
sys.path.append('.')  # Добавляем текущую директорию в путь поиска
sys.modules['src.scripts_for_bash_with_inheritance.app_logger'] = MagicMock()
from src.scripts_for_bash_with_inheritance.reference_compass import ReferenceCompass, headers_eng, list_join_columns


# Настраиваем фикстуры и моки
@pytest.fixture
def mock_env_vars():
    with mock.patch.dict(os.environ, {
        'HOST': 'test_host',
        'DATABASE': 'test_db',
        'USERNAME_DB': 'test_user',
        'PASSWORD': 'test_password'
    }):
        yield


@pytest.fixture
def reference_compass():
    return ReferenceCompass('test_input.xlsx', 'test_output/')


@pytest.fixture
def mock_client():
    mock_client = mock.MagicMock()
    mock_client.query.return_value.result_rows = []
    return mock_client


@pytest.fixture
def mock_get_client(mock_client):
    with mock.patch('src.scripts_for_bash_with_inheritance.reference_compass.get_client', return_value=mock_client):
        yield mock_client


@pytest.fixture
def mock_requests_post():
    with mock.patch('requests.post') as mock_post:
        yield mock_post




def test_convert_format_date(reference_compass):
    """Test date format conversion"""
    # Тестируем разные форматы дат
    assert reference_compass.convert_format_date("03/15/24") == "2024-03-15"
    assert reference_compass.convert_format_date("15.03.2024") == "2024-03-15"
    assert reference_compass.convert_format_date("2024-03-15 12:30:45") == "2024-03-15"
    assert reference_compass.convert_format_date("03/15/2024") == "2024-03-15"
    assert reference_compass.convert_format_date("15Mar2024") == "2024-03-15"

    # Некорректный формат должен вернуть None
    assert reference_compass.convert_format_date("некорректная дата") is None


def test_leave_largest_data_with_dupl_inn(reference_compass):
    """Test filtering duplicates by INN"""
    # Создаем тестовые данные с дублирующимися ИНН
    data = [
        {"inn": "1234567890", "company_name": "Company A", "address": "Address A"},
        {"inn": "1234567890", "company_name": "Company A", "address": None},
        {"inn": "9876543210", "company_name": "Company B", "address": "Address B"}
    ]

    # Запускаем тестируемую функцию
    result = reference_compass.leave_largest_data_with_dupl_inn(data)

    # Проверяем результат
    assert len(result) == 2
    assert result[0]["inn"] == "1234567890"
    assert result[0]["address"] == "Address A"
    assert result[1]["inn"] == "9876543210"


def test_add_new_columns(reference_compass):
    """Test adding new columns to data dictionary"""
    # Создаем тестовый словарь данных
    data = {"inn": "1234567890", "company_name": "Test Company"}

    # Запускаем тестируемую функцию
    with mock.patch('src.scripts_for_bash_with_inheritance.reference_compass.datetime') as mock_datetime:
        mock_datetime.now.return_value.strftime.return_value = "2024-03-17 10:00:00"
        reference_compass.add_new_columns(data)

    # Проверяем результат
    assert data["original_file_name"] == "test_input.xlsx"
    assert data["original_file_parsed_on"] == "2024-03-17 10:00:00"
    assert data["last_updated"] == "2024-03-17 10:00:00"
    assert data["dadata_branch_name"] == ""
    assert data["dadata_branch_address"] == ""
    assert data["dadata_branch_region"] == ""


def test_get_data_from_service_inn(reference_compass, mock_requests_post):
    """Test getting data from service_inn API"""
    # Настраиваем мок для запроса к API
    mock_response = mock.MagicMock()
    mock_response.json.return_value = [
        [
            {
                "data": {
                    "name": {"full": "Test Company"},
                    "opf": {"short": "ООО"},
                    "branch_type": "MAIN",
                    "state": {"status": "ACTIVE", "registration_date": 1577836800000, "liquidation_date": None},
                    "address": {
                        "unrestricted_value": "Test Address",
                        "data": {"region_with_type": "Test Region", "federal_district": "Test District",
                                 "city": "Test City", "geo_lat": "55.123", "geo_lon": "37.456"}
                    },
                    "okpo": "12345678",
                    "okved": "Test OKVED",
                    "kpp": "123456789",
                    "type": "LEGAL"
                },
                "value": "Test Value"
            }
        ],
        False
    ]
    mock_response.raise_for_status.return_value = None
    mock_requests_post.return_value = mock_response

    # Создаем тестовый словарь данных с необходимыми полями
    data = {
        "inn": "1234567890",
        "company_name": "Test Company",
        "original_file_name": "test_input.xlsx",
        "original_file_parsed_on": "2024-03-17 10:00:00"
    }

    # Мокаем метод save_to_csv, чтобы избежать ошибки
    with mock.patch.object(reference_compass, 'save_to_csv'):
        # Запускаем тестируемую функцию
        reference_compass.get_data_from_service_inn(data, 2)

        # Проверяем результат
        mock_requests_post.assert_called_once_with("http://service_inn:8003", json={"inn": "1234567890"})
        assert data["dadata_company_name"] == "ООО Test Company"
        assert data["dadata_address"] == "Test Address"
        assert data["dadata_region"] == "Test Region"
        assert data["dadata_status"] == "ACTIVE"
        assert data["dadata_registration_date"] == "2020-01-01"


def test_get_data_from_service_inn_exception(reference_compass, mock_requests_post):
    """Test handling exception when getting data from service_inn API"""
    # Настраиваем мок для запроса к API, который вызывает исключение
    mock_requests_post.side_effect = RequestException("API error")

    # Создаем тестовый словарь данных
    data = {"inn": "1234567890", "company_name": "Test Company"}

    # Запускаем тестируемую функцию (не должна вызывать исключение)
    reference_compass.get_data_from_service_inn(data, 2)

    # Проверяем, что запрос был вызван с правильными параметрами
    mock_requests_post.assert_called_once_with("http://service_inn:8003", json={"inn": "1234567890"})


def test_handle_raw_data(reference_compass):
    """Test handling raw data"""
    # Создаем тестовые данные
    parsed_data = [
        {"inn": "1234567890", "company_name": "Company A", "registration_date": "15.03.2024",
         "revenue_at_upload_date_thousand_rubles": "1000", "employees_number_at_upload_date": "50",
         "net_profit_or_loss_at_upload_date_thousand_rubles": "500"},
        {"inn": "invalid_inn", "company_name": "Invalid Company"}
    ]

    # Мокаем методы, которые вызываются внутри handle_raw_data
    with mock.patch.object(reference_compass, 'get_data_from_cache') as mock_get_data_from_cache, \
            mock.patch.object(reference_compass, 'save_to_csv') as mock_save_to_csv, \
            mock.patch('src.scripts_for_bash_with_inheritance.reference_compass.is_valid', side_effect=lambda x: x == "1234567890"):
        # Запускаем тестируемую функцию
        reference_compass.handle_raw_data(parsed_data)

        # Проверяем результат
        assert len(parsed_data) == 1
        assert parsed_data[0]["inn"] == "1234567890"
        assert parsed_data[0]["registration_date"] == "2024-03-15"
        assert parsed_data[0]["revenue_at_upload_date_thousand_rubles"] == 1000
        assert parsed_data[0]["employees_number_at_upload_date"] == 50
        assert parsed_data[0]["net_profit_or_loss_at_upload_date_thousand_rubles"] == 500

        # Проверяем, что методы были вызваны с правильными параметрами
        mock_get_data_from_cache.assert_called_once()
        mock_save_to_csv.assert_called_once()


def test_parse_xlsx(reference_compass):
    """Test parsing XLSX file"""
    # Создаем тестовый Excel-файл
    wb = Workbook()
    ws = wb.active

    # Заполняем шапку таблицы
    ws.cell(row=1, column=1).value = "ИНН"
    ws.cell(row=1, column=2).value = "Наименование"
    ws.cell(row=1, column=3).value = "Электронная почта"

    # Заполняем данные
    ws.cell(row=2, column=1).value = "1234567890"
    ws.cell(row=2, column=2).value = "Test Company"
    ws.cell(row=2, column=3).value = "test@example.com"

    # Мокаем метод для получения значений из ячеек
    with mock.patch.object(reference_compass, 'get_column_eng') as mock_get_column_eng, \
            mock.patch.object(reference_compass, 'get_value_from_cell') as mock_get_value_from_cell:

        # Мокаем поведение метода get_column_eng
        def side_effect_get_column_eng(column, dict_header):
            if column[0].value == "ИНН":
                dict_header["A"] = ("ИНН", "inn")
            elif column[0].value == "Наименование":
                dict_header["B"] = ("Наименование", "company_name")
            elif column[0].value == "Электронная почта":
                dict_header["C"] = ("Электронная почта", "email")

        mock_get_column_eng.side_effect = side_effect_get_column_eng

        # Мокаем поведение метода get_value_from_cell
        def side_effect_get_value_from_cell(column, dict_header, dict_columns):
            if column[0].row == 2:
                dict_columns["inn"] = "1234567890"
                dict_columns["company_name"] = "Test Company"
                dict_columns["email"] = "test@example.com"

        mock_get_value_from_cell.side_effect = side_effect_get_value_from_cell

        # Запускаем тестируемую функцию
        parsed_data = []
        reference_compass.parse_xlsx(ws, parsed_data)

        # Проверяем результат
        assert len(parsed_data) == 1
        assert parsed_data[0]["inn"] == "1234567890"
        assert parsed_data[0]["company_name"] == "Test Company"
        assert parsed_data[0]["email"] == "test@example.com"


def test_main(reference_compass):
    """Test main function"""
    # Мокаем методы, которые вызываются внутри main
    with mock.patch('src.scripts_for_bash_with_inheritance.reference_compass.load_workbook') as mock_load_workbook, \
            mock.patch.object(reference_compass, 'parse_xlsx') as mock_parse_xlsx, \
            mock.patch.object(reference_compass, 'handle_raw_data') as mock_handle_raw_data, \
            mock.patch.object(reference_compass, 'leave_largest_data_with_dupl_inn',
                              return_value=[{"inn": "1234567890"}]) as mock_leave_largest_data, \
            mock.patch.object(reference_compass, 'change_data_in_db') as mock_change_data_in_db, \
            mock.patch.object(reference_compass, 'write_to_json') as mock_write_to_json:
        # Настраиваем мок для load_workbook
        mock_wb = mock.MagicMock()
        mock_wb.sheetnames = ["Sheet1"]
        mock_wb.__getitem__.return_value = "worksheet"
        mock_load_workbook.return_value = mock_wb

        # Запускаем тестируемую функцию
        reference_compass.main()

        # Проверяем, что методы были вызваны с правильными параметрами
        mock_load_workbook.assert_called_once_with('test_input.xlsx')
        mock_parse_xlsx.assert_called_once_with("worksheet", [])
        mock_handle_raw_data.assert_called_once()
        mock_leave_largest_data.assert_called_once()
        mock_change_data_in_db.assert_called_once_with([{"inn": "1234567890"}])
        mock_write_to_json.assert_called_once_with([{"inn": "1234567890"}])
